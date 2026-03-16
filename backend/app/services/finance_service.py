"""
Finance Service — querying, anomaly detection, P&L, cashflow, reconciliation, and export.

Operates on finance_document signals stored in DynamoDB. Provides:
- Duplicate and statistical-outlier anomaly detection
- Review queue management (approve / reject / edit)
- Cashflow aggregation by period (daily, weekly, monthly)
- P&L computation with full Nigerian tax breakdown
- Bank reconciliation against uploaded CSV/Excel statements
- CSV and XLSX export of finance documents
"""

import csv
import io
import logging
import statistics
from collections import defaultdict
from datetime import datetime, timedelta
from difflib import SequenceMatcher
from typing import Dict, Any, List, Optional

import openpyxl

from app.models.finance_document import DocumentFlag
from app.services.ddb_service import DynamoDBService, get_ddb_service

logger = logging.getLogger(__name__)


class FinanceService:
    """Service for finance document operations: querying, anomaly detection, review workflow."""

    def __init__(self, ddb_service: DynamoDBService):
        self.ddb_service = ddb_service

    def get_finance_documents(self, org_id: str) -> List[Dict[str, Any]]:
        """Query all finance documents for an organization.

        Fetches signals from DynamoDB and filters to signal_type='finance_document'.
        """
        signals = self.ddb_service.query_signals(org_id)
        return [s for s in signals if s.get("signal_type") == "finance_document"]

    def detect_anomalies(
        self, org_id: str, document: Dict[str, Any]
    ) -> List[DocumentFlag]:
        """Detect duplicate and statistical-outlier anomalies for a document.

        Checks against existing finance documents for the same org:
        - Duplicate: same vendor_name + amount + document_date in content.
        - Statistical outlier: amount > 3σ from mean for the same vendor.

        Returns a list of DocumentFlag objects (may be empty).
        """
        existing_docs = self.get_finance_documents(org_id)
        content = document.get("content", {})
        vendor = content.get("vendor_name", "")
        amount = content.get("amount")
        doc_date = content.get("document_date")

        flags: List[DocumentFlag] = []

        # --- Duplicate detection ---
        for existing in existing_docs:
            ec = existing.get("content", {})
            if (
                ec.get("vendor_name") == vendor
                and ec.get("amount") == amount
                and ec.get("document_date") == doc_date
            ):
                flags.append(
                    DocumentFlag(
                        flag_type="duplicate",
                        flag_reason=(
                            f"Duplicate: same vendor '{vendor}', amount {amount}, "
                            f"and date {doc_date} found in existing document"
                        ),
                    )
                )
                break  # one duplicate flag is sufficient

        # --- Statistical outlier detection ---
        same_vendor_amounts: List[float] = []
        for existing in existing_docs:
            ec = existing.get("content", {})
            if ec.get("vendor_name") == vendor:
                existing_amount = ec.get("amount")
                if existing_amount is not None:
                    try:
                        same_vendor_amounts.append(float(existing_amount))
                    except (TypeError, ValueError):
                        continue

        if len(same_vendor_amounts) >= 2 and amount is not None:
            try:
                mean = statistics.mean(same_vendor_amounts)
                stdev = statistics.stdev(same_vendor_amounts)
                if stdev > 0 and abs(float(amount) - mean) > 3 * stdev:
                    flags.append(
                        DocumentFlag(
                            flag_type="anomaly",
                            flag_reason=(
                                f"Statistical outlier: amount {amount} is more than "
                                f"3 standard deviations from the mean ({mean:.2f}) "
                                f"for vendor '{vendor}'"
                            ),
                        )
                    )
            except (TypeError, ValueError):
                pass

        return flags

    def get_review_queue(self, org_id: str) -> List[Dict[str, Any]]:
        """Return finance documents whose processing_status is 'needs_review'."""
        docs = self.get_finance_documents(org_id)
        return [
            d for d in docs
            if d.get("content", {}).get("processing_status") == "needs_review"
        ]

    def update_review_status(
        self,
        org_id: str,
        signal_id: str,
        action: str,
        edits: Optional[Dict[str, Any]] = None,
    ) -> None:
        """Approve, reject, or edit a finance document in the review queue.

        Actions:
        - 'approve': sets processing_status to 'approved'
        - 'reject': sets processing_status to 'rejected'
        - 'edit': applies *edits* to the content fields and sets processing_status to 'approved'
        """
        signal = self.ddb_service.get_signal(org_id, signal_id)
        if signal is None:
            raise ValueError(f"Signal {signal_id} not found for org {org_id}")

        content = signal.get("content", {})

        if action == "approve":
            content["processing_status"] = "approved"
        elif action == "reject":
            content["processing_status"] = "rejected"
        elif action == "edit":
            if edits:
                content.update(edits)
            content["processing_status"] = "approved"
        else:
            raise ValueError(f"Invalid action '{action}'. Must be 'approve', 'reject', or 'edit'.")

        self.ddb_service.update_signal(org_id, signal_id, {"content": content})

    # ==================== Cashflow & P&L ====================

    def _filter_active_docs(
        self, org_id: str, start_date: Optional[str] = None, end_date: Optional[str] = None
    ) -> List[Dict[str, Any]]:
        """Return finance documents with status 'approved' or 'processed', optionally filtered by date range."""
        docs = self.get_finance_documents(org_id)
        active = [
            d for d in docs
            if d.get("content", {}).get("processing_status") in ("approved", "processed")
        ]
        if not start_date and not end_date:
            return active

        filtered = []
        for d in active:
            doc_date_str = d.get("content", {}).get("document_date")
            if not doc_date_str:
                continue
            try:
                doc_date = datetime.fromisoformat(doc_date_str) if isinstance(doc_date_str, str) else doc_date_str
                if start_date:
                    sd = datetime.fromisoformat(start_date) if isinstance(start_date, str) else start_date
                    if doc_date < sd:
                        continue
                if end_date:
                    ed = datetime.fromisoformat(end_date) if isinstance(end_date, str) else end_date
                    if doc_date > ed:
                        continue
                filtered.append(d)
            except (ValueError, TypeError):
                continue
        return filtered

    @staticmethod
    def _period_key(dt: datetime, period: str) -> str:
        """Return a grouping key string for the given period granularity."""
        if period == "daily":
            return dt.strftime("%Y-%m-%d")
        elif period == "weekly":
            iso = dt.isocalendar()
            return f"{iso[0]}-W{iso[1]:02d}"
        else:  # monthly
            return dt.strftime("%Y-%m")

    def get_cashflow(
        self,
        org_id: str,
        period: str = "monthly",
        start_date: Optional[str] = None,
        end_date: Optional[str] = None,
    ) -> List[Dict[str, Any]]:
        """Aggregate revenue/expense totals grouped by period.

        Returns a list of dicts: [{period, revenue, expenses}, ...]
        """
        docs = self._filter_active_docs(org_id, start_date, end_date)
        buckets: Dict[str, Dict[str, float]] = defaultdict(lambda: {"revenue": 0.0, "expenses": 0.0})

        for d in docs:
            c = d.get("content", {})
            doc_date_str = c.get("document_date")
            if not doc_date_str:
                continue
            try:
                dt = datetime.fromisoformat(doc_date_str) if isinstance(doc_date_str, str) else doc_date_str
            except (ValueError, TypeError):
                continue

            key = self._period_key(dt, period)
            amount = float(c.get("amount", 0))
            category = c.get("category", "")
            if category in ("payment", "credit_note", "revenue"):
                buckets[key]["revenue"] += amount
            elif category in ("invoice", "receipt", "expense", "purchase_order"):
                buckets[key]["expenses"] += amount

        result = [
            {"period": k, "revenue": round(v["revenue"], 2), "expenses": round(v["expenses"], 2)}
            for k, v in sorted(buckets.items())
        ]
        return result

    def get_pnl(
        self,
        org_id: str,
        start_date: Optional[str] = None,
        end_date: Optional[str] = None,
    ) -> Dict[str, Any]:
        """Compute P&L summary with full Nigerian tax breakdown (VAT, WHT, CIT, PAYE, customs)."""
        docs = self._filter_active_docs(org_id, start_date, end_date)

        total_revenue = 0.0
        total_expenses = 0.0
        total_vat_collected = 0.0
        total_vat_paid = 0.0
        total_wht = 0.0
        total_cit = 0.0
        total_paye = 0.0
        total_customs = 0.0
        vendor_breakdown: Dict[str, Dict[str, float]] = defaultdict(lambda: {"revenue": 0.0, "expenses": 0.0})

        for d in docs:
            c = d.get("content", {})
            amount = float(c.get("amount", 0))
            category = c.get("category", "")
            vendor = c.get("vendor_name", "Unknown")
            vat_amount = float(c.get("vat_amount", 0) or 0)
            wht_amount = float(c.get("wht_amount", 0) or 0)
            cit_amount = float(c.get("cit_amount", 0) or 0)
            paye_amount = float(c.get("paye_amount", 0) or 0)
            customs_levy = float(c.get("customs_levy", 0) or 0)

            if category in ("payment", "credit_note", "revenue"):
                total_revenue += amount
                vendor_breakdown[vendor]["revenue"] += amount
                total_vat_collected += vat_amount
            elif category in ("invoice", "receipt", "expense", "purchase_order"):
                total_expenses += amount
                vendor_breakdown[vendor]["expenses"] += amount
                total_vat_paid += vat_amount

            # These taxes apply regardless of category
            total_wht += wht_amount
            total_cit += cit_amount
            total_paye += paye_amount
            total_customs += customs_levy

        return {
            "total_revenue": round(total_revenue, 2),
            "total_expenses": round(total_expenses, 2),
            "net_profit": round(total_revenue - total_expenses, 2),
            "by_vendor": {
                v: {"revenue": round(d["revenue"], 2), "expenses": round(d["expenses"], 2)}
                for v, d in sorted(vendor_breakdown.items())
            },
            "vat_summary": {
                "total_vat_collected": round(total_vat_collected, 2),
                "total_vat_paid": round(total_vat_paid, 2),
            },
            "tax_summary": {
                "vat_collected": round(total_vat_collected, 2),
                "vat_paid": round(total_vat_paid, 2),
                "vat_net": round(total_vat_collected - total_vat_paid, 2),
                "withholding_tax": round(total_wht, 2),
                "corporate_income_tax": round(total_cit, 2),
                "paye_payroll": round(total_paye, 2),
                "customs_levy": round(total_customs, 2),
                "total_tax_burden": round(
                    total_vat_paid + total_wht + total_cit + total_paye + total_customs, 2
                ),
            },
        }

    # ==================== Reconciliation & CSV Export ====================

    def reconcile(self, org_id: str, bank_csv_content: str) -> Dict[str, Any]:
        """Match bank CSV transactions against finance documents.

        Matching uses three criteria (all must pass):
        - Amount within 1% tolerance
        - Date within 3 calendar days
        - Vendor name similarity > 0.6 (SequenceMatcher ratio)

        Returns matched pairs, unmatched transactions, and unmatched documents.
        """
        # Parse bank CSV
        reader = csv.DictReader(io.StringIO(bank_csv_content))
        bank_txns = []
        for row in reader:
            try:
                bank_txns.append({
                    "date": row.get("date", ""),
                    "description": row.get("description", ""),
                    "amount": float(row.get("amount", 0)),
                    "matched": False,
                    "matched_document_id": None,
                })
            except (ValueError, TypeError):
                continue

        docs = self._filter_active_docs(org_id)
        doc_matched = set()
        matched_pairs = []

        for txn in bank_txns:
            for d in docs:
                c = d.get("content", {})
                doc_id = c.get("document_id", d.get("signal_id", ""))
                if doc_id in doc_matched:
                    continue

                # Amount tolerance: 1%
                doc_amount = float(c.get("amount", 0))
                if doc_amount == 0:
                    continue
                if abs(txn["amount"] - doc_amount) / doc_amount > 0.01:
                    continue

                # Date tolerance: 3 days
                try:
                    txn_date = datetime.fromisoformat(txn["date"]) if txn["date"] else None
                    doc_date_str = c.get("document_date", "")
                    doc_date = datetime.fromisoformat(doc_date_str) if isinstance(doc_date_str, str) else doc_date_str
                    if txn_date and doc_date and abs((txn_date - doc_date).days) > 3:
                        continue
                except (ValueError, TypeError):
                    continue

                # Vendor similarity
                vendor = c.get("vendor_name", "")
                similarity = SequenceMatcher(None, txn["description"].lower(), vendor.lower()).ratio()
                if similarity < 0.6:
                    continue

                # Match found
                txn["matched"] = True
                txn["matched_document_id"] = doc_id
                doc_matched.add(doc_id)
                matched_pairs.append({"transaction": txn, "document_id": doc_id})
                break

        unmatched_txns = [t for t in bank_txns if not t["matched"]]
        unmatched_docs = [
            d for d in docs
            if (d.get("content", {}).get("document_id", d.get("signal_id", ""))) not in doc_matched
        ]

        return {
            "matched": matched_pairs,
            "unmatched_transactions": unmatched_txns,
            "unmatched_documents": unmatched_docs,
            "summary": {
                "total_transactions": len(bank_txns),
                "matched_count": len(matched_pairs),
                "unmatched_transaction_count": len(unmatched_txns),
                "unmatched_document_count": len(unmatched_docs),
            },
        }

    # ---- shared helpers for export ----

    _EXPORT_HEADERS = [
        "document_id", "vendor_name", "amount", "currency",
        "vat_amount", "vat_rate", "wht_amount", "cit_amount",
        "paye_amount", "customs_levy", "document_date", "category",
        "confidence_score", "processing_status",
    ]

    def _export_rows(
        self,
        org_id: str,
        start_date: Optional[str] = None,
        end_date: Optional[str] = None,
        category: Optional[str] = None,
    ) -> List[List[Any]]:
        """Return header + data rows for export."""
        docs = self._filter_active_docs(org_id, start_date, end_date)
        if category:
            docs = [d for d in docs if d.get("content", {}).get("category") == category]

        rows: List[List[Any]] = []
        for d in docs:
            c = d.get("content", {})
            rows.append([
                c.get("document_id", d.get("signal_id", "")),
                c.get("vendor_name", ""),
                c.get("amount", ""),
                c.get("currency", ""),
                c.get("vat_amount", ""),
                c.get("vat_rate", ""),
                c.get("wht_amount", ""),
                c.get("cit_amount", ""),
                c.get("paye_amount", ""),
                c.get("customs_levy", ""),
                c.get("document_date", ""),
                c.get("category", ""),
                c.get("confidence_score", ""),
                c.get("processing_status", ""),
            ])
        return rows

    def export_csv(
        self,
        org_id: str,
        start_date: Optional[str] = None,
        end_date: Optional[str] = None,
        category: Optional[str] = None,
    ) -> str:
        """Generate CSV string of finance documents."""
        rows = self._export_rows(org_id, start_date, end_date, category)
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(self._EXPORT_HEADERS)
        writer.writerows(rows)
        return output.getvalue()

    def export_xlsx(
        self,
        org_id: str,
        start_date: Optional[str] = None,
        end_date: Optional[str] = None,
        category: Optional[str] = None,
    ) -> bytes:
        """Generate Excel (.xlsx) bytes of finance documents."""
        rows = self._export_rows(org_id, start_date, end_date, category)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Finance Export"
        ws.append(self._EXPORT_HEADERS)
        for row in rows:
            ws.append(row)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ---- spreadsheet / CSV import helpers ----

    @staticmethod
    def parse_spreadsheet_to_csv(file_content: bytes, content_type: str, filename: str) -> str:
        """Convert uploaded CSV or Excel bytes into a normalised CSV string.

        Supports text/csv, .xlsx, and .xls formats. Raises ValueError for unsupported types.
        """
        ct = (content_type or "").lower()
        fn = (filename or "").lower()

        # CSV
        if "csv" in ct or fn.endswith(".csv"):
            return file_content.decode("utf-8")

        # Excel
        if fn.endswith(".xlsx") or fn.endswith(".xls") or "spreadsheet" in ct or "ms-excel" in ct:
            wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
            ws = wb.active
            output = io.StringIO()
            writer = csv.writer(output)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(row)
            wb.close()
            return output.getvalue()

        raise ValueError(f"Unsupported file type: {content_type} ({filename})")


import threading


_finance_service: Optional[FinanceService] = None
_fs_lock = threading.Lock()


def get_finance_service() -> FinanceService:
    """Get singleton FinanceService instance (thread-safe)"""
    global _finance_service
    if _finance_service is None:
        with _fs_lock:
            if _finance_service is None:
                _finance_service = FinanceService(get_ddb_service())
    return _finance_service
